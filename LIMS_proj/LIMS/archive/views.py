from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import HttpResponseRedirect, render, get_object_or_404, render_to_response, redirect
from django.forms import modelformset_factory, inlineformset_factory
from django import forms
import django_excel as excel
from django.core.files.storage import FileSystemStorage
from django.template import RequestContext
import pyexcel as pe
from openpyxl import load_workbook
import xlrd
import xlwt
import csv
import codecs
import sys
from string import ascii_uppercase
import copy
from project.models import Project
from .forms import IndividualForm, SampleForm, SampleMultiForm, UploadFileForm, FullSampleForm, FreezerForm, BoxForm
from .models import Individual, Sample, Freezer, FreezerPos, Box, BoxPos
from .tables import IndividualTable, SampleTable, FullSampleTable, FreezerTable, BoxTable
from .filters import SampleFilter, IndividualFilter, FreezerFilter
from .resources import SampleResource, FullSampleResource


### global

# button variables
# add button variables
add_sample_single_button_context = {
    "full_url": 'add_full_sample_single',
    "full_name": 'New Everything',
    "parent_url": 'add_individual_single',
    "parent_name": 'New Individual',
    "child_url": 'add_sample_single',
    "child_name": 'New Sample',
}

add_sample_multiple_button_context = {
    "full_url": 'add_full_sample_multiple',
    "full_name": 'New Everything',
    "parent_url": 'add_individual_multiple',
    "parent_name": 'New Individual',
    "child_url": 'add_sample_multiple',
    "child_name": 'New Sample',
}

add_freezer_button_context = {
    "full_url": 'add_full_sample_single',
    "full_name": 'New Everything',
    "parent_url": 'add_freezer',
    "parent_name": 'New Freezer',
    "child_url": 'add_box',
    "child_name": 'New Box',
}

validate_samples_button_context = {
    "full_url": 'add_full_sample_single',
    "full_name": 'New Everything',
    "parent_url": 'validate_individual_multiple',
    "parent_name": 'Individuals',
    "child_url": 'validate_sample_multiple',
    "child_name": 'Samples',
}

### home
def archive(request):
    return render(request, 'individual.html', {})


### individuals/samples
# add single
def add_full_sample_single(request):
    title = "Enter Individual and Sample"
    single_url = 'add_full_sample_single'
    multi_url = 'add_full_sample_multiple'
    p_form = IndividualForm(prefix="ind")
    c_form = SampleMultiForm(prefix="samp")
    table = FullSampleTable(Sample.objects.all())

    if request.POST:
        p_form = IndividualForm(request.POST, prefix="ind")
        c_form = SampleMultiForm(request.POST, prefix="samp")
        if p_form.is_valid() and c_form.is_valid():
            sample = c_form.save(commit=False)
            sample.individual = p_form.save()
            sample.save()

    context = {
        "button_type": 'buttons_1.html',
        "title": title,
        "single_url": single_url,
        "multi_url": multi_url,
        "form1": p_form,
        "form2": c_form,
        "table": table,
    }
    context.update(add_sample_single_button_context)
    return render(request, 'add_single.html', context)


def add_sample_single(request):
    title = 'Enter Sample'
    single_url = 'add_sample_single'
    multi_url = 'add_sample_multiple'
    form = SampleForm(request.POST or None)
    form2 = 'n/a'
    table = SampleTable(Sample.objects.all())

    if form.is_valid():
        entered_sample = form.save()
        pk = entered_sample.pk
        return HttpResponseRedirect('/archive/edit_sample_single/%s' % pk)

    context = {
        "button_type": 'buttons_1.html',
        "title": title,
        "single_url": single_url,
        "multi_url": multi_url,
        "form1": form,
        "form2": form2,
        "table": table,
    }
    context.update(add_sample_single_button_context)
    return render(request, 'add_single.html', context)


def add_individual_single(request):
    title = 'Enter Individual'
    single_url = 'add_individual_single'
    multi_url = 'add_individual_multiple'
    form = IndividualForm(request.POST or None)
    form2 = 'n/a'
    table = IndividualTable(Individual.objects.all())

    if form.is_valid():
        instance = form.save(commit=False)
        full_name = form.cleaned_data.get("full_name")
        if not full_name:
            full_name = "New full name"

        instance.full_name = full_name
        instance.save()

    context = {
        "button_type": 'buttons_1.html',
        "single_url": single_url,
        "multi_url": multi_url,
        "title": title,
        "form1": form,
        "form2": form2,
        "table": table,
    }
    context.update(add_sample_single_button_context)
    return render(request, 'add_single.html', context)


# add multiple
def add_full_sample_multiple(request):
    title = 'Enter Individual and Sample Data'
    single_url = 'add_sample_single'
    multi_url = 'add_sample_multiple'
    extras = 5
    p_data = []
    c_data = []

    if "quantity" in request.POST:
        extras = int(request.POST['quantity'])

    if "export_btn" in request.POST:
        # from https://simpleisbetterthancomplex.com/tutorial/2016/07/29/how-to-export-to-excel.html
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="full_data.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Sample')
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = ['indv_name', 'project', 'species', 'family_id', 'relationship',
                   'karyotype', 'other_genetic_info', 'gender', 'year_of_birth',
                   'sample_name', 'type', 'o_tissue', 'conc', 'vol']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        wb.save(response)
        return response

    if "upload_btn" in request.POST:
        upload_form = UploadFileForm(request.POST, request.FILES)
        if upload_form.is_valid():
            filehandle = request.FILES['myfile']
            data = filehandle.get_records()
            count = 0
            for record in data:
                count = count + 1
                #find Project IDs
                project_name = record['project']
                project_record = Project.objects.get_or_create(name=project_name)
                project = project_record[0]
                record['project'] = project.id

                #make p_record
                p_record = copy.deepcopy(record)
                p_record['name'] = p_record['indv_name']
                p_data.append(p_record)

                #make c_record
                c_record = copy.deepcopy(record)
                c_record['name'] = c_record['sample_name']
                c_data.append(c_record)
            extras = count

    p_IndividualFormSet = modelformset_factory(Individual, form=IndividualForm, extra=extras)
    p_qset = Individual.objects.none()
    p_formset = p_IndividualFormSet(queryset=p_qset, initial=p_data)
    c_SampleFormSet = modelformset_factory(Sample, form=SampleMultiForm, extra=extras)
    c_qset = Sample.objects.none()
    c_formset = c_SampleFormSet(queryset=c_qset, initial=c_data)
    table = FullSampleTable(Sample.objects.all())
    upload_form = UploadFileForm()

    if "save_btn" in request.POST:
        p_formset = p_IndividualFormSet(request.POST)
        c_formset = c_SampleFormSet(request.POST)
        p_ids = list()
        if p_formset.is_valid and c_formset.is_valid:
            for form in p_formset:
                if form.is_valid():
                    indv_name = form.cleaned_data.get('name')
                    qs = Individual.objects.filter().values_list('name', flat=True)
                    if indv_name in qs:
                        indv_record = Individual.objects.get(name=indv_name)
                        p_ids.append(indv_record.id)
                    else:
                        indv_record = form.save()
                        p_ids.append(indv_record.id)
            p_id_num = -1
            for form in c_formset:
                p_id_num = p_id_num + 1
                if form.is_valid():
                    c_form = form.save(commit=False)
                    c_form.individual_id = p_ids[p_id_num]
                    c_form.save()
        return HttpResponseRedirect('/archive/validate_sample_multiple')


    context = {
        "button_type": 'buttons_1.html',
        "title": title,
        "single_url": single_url,
        "multi_url": multi_url,
        "p_formset": p_formset,
        "c_formset": c_formset,
        "table": table,
        "extras": extras,
        "upload_form": upload_form,
    }
    context.update(add_sample_multiple_button_context)
    return render(request, 'add_multiple.html', context)


def add_sample_multiple(request):
    title = 'Enter Multiple Samples'
    single_url = 'add_sample_single'
    multi_url = 'add_sample_multiple'
    extras = 5
    pks = list()
    qset = Sample.objects.none()
    data = ()
    upload_form = UploadFileForm()

    if "quantity" in request.POST:
        extras = int(request.POST['quantity'])

    if "export_btn" in request.POST:
        # from https://simpleisbetterthancomplex.com/tutorial/2016/07/29/how-to-export-to-excel.html
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="samples.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Sample')
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = ['individual', 'sample_name', 'type', 'o_tissue', 'conc', 'vol']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        wb.save(response)
        return response

    if "upload_btn" in request.POST:
        upload_form = UploadFileForm(request.POST, request.FILES)
        if upload_form.is_valid():
            filehandle = request.FILES['myfile']
            data = filehandle.get_records()
            count = 0
            for record in data:
                count = count + 1
                indv_name = record['individual']
                indv_record = Individual.objects.get_or_create(name=indv_name)
                indv = indv_record[0]
                record['individual'] = indv.id
            extras = count

    SampleFormSet = modelformset_factory(Sample, form=SampleForm, extra=extras)
    p_formset = SampleFormSet(queryset=qset, initial=data)
    c_formset = 'n/a'
    table = SampleTable(Sample.objects.all())

    if "save_btn" in request.POST:
        p_formset = SampleFormSet(request.POST)
        for form in p_formset:
            if form.is_valid():
                entered_sample = form.save()
                pk = entered_sample.pk
                pks.append(pk)
        return HttpResponseRedirect('/archive/validate_sample_multiple')

    context = {
        "button_type": 'buttons_1.html',
        "title": title,
        "single_url": single_url,
        "multi_url": multi_url,
        "p_formset": p_formset,
        "c_formset": c_formset,
        "table": table,
        "extras": extras,
        "upload_form": upload_form,
    }
    context.update(add_sample_multiple_button_context)
    return render(request, 'add_multiple.html', context)


def add_individual_multiple(request):
    title = 'Enter Multiple Individuals'
    single_url = 'add_sample_single'
    multi_url = 'add_sample_multiple'
    extras = 5
    qset = Individual.objects.none()
    data = ()
    upload_form = UploadFileForm()

    if "quantity" in request.POST:
        extras = int(request.POST['quantity'])

    if "export_btn" in request.POST:
        # from https://simpleisbetterthancomplex.com/tutorial/2016/07/29/how-to-export-to-excel.html
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="example.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Sample')
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = ['name', 'alt_name1', 'project', 'species', 'family_id', 'relationship',
                   'karyotype', 'other_genetic_info', 'gender', 'year_of_birth']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        wb.save(response)
        return response

    if "upload_btn" in request.POST:
        upload_form = UploadFileForm(request.POST, request.FILES)
        if upload_form.is_valid():
            filehandle = request.FILES['myfile']
            data = filehandle.get_records()
            count = 0
            for record in data:
                count = count + 1
                project_name = record['project']
                project_record = Project.objects.get_or_create(name=project_name)
                project = project_record[0]
                record['project'] = project.id
            extras = count


    IndividualFormSet = modelformset_factory(Individual, form=IndividualForm, extra=extras)
    p_formset = IndividualFormSet(queryset=qset, initial=data)
    c_formset = 'n/a'
    table = IndividualTable(Individual.objects.all())

    if "save_btn" in request.POST:
        p_formset = IndividualFormSet(request.POST)
        for form in p_formset:
            if form.is_valid():
                form.save()
        return HttpResponseRedirect('/archive/validate_individual_multiple')

    context = {
        "button_type": 'buttons_1.html',
        "title": title,
        "single_url": single_url,
        "multi_url": multi_url,
        "p_formset": p_formset,
        "c_formset": c_formset,
        "table": table,
        "extras": extras,
        "upload_form": upload_form,
    }
    context.update(add_sample_multiple_button_context)
    return render(request, 'add_multiple.html', context)


# edit
def edit_sample_single(request, id):
    table = SampleTable(Sample.objects.all())
    sample = Sample.objects.get(pk=id)
    form = SampleForm(instance=sample)
    add_url = 'add_sample_single'
    delete_url = 'delete_sample'

    if request.method == "POST":
        if "del_btn" in request.POST:
            return HttpResponseRedirect('/archive/delete_sample/%s' % id)

        elif "edit_btn" in request.POST:
            title = "Sample Edited. Edit or Add Another?"
            instance = get_object_or_404(Sample, id=id)
            form = SampleForm(request.POST or None, instance=instance)
            if form.is_valid():
                form.save()

        elif "add_btn" in request.POST:
            return HttpResponseRedirect('/archive/add_sample_single/%s' % id)
    elif id:
        title = "Sample Added. Edit or Add Another?"
        table = SampleTable(Sample.objects.all())

    else:
        title = "Sample Failed to Add. Retry?."

    context = {
        "title": title,
        "form": form,
        "table": table,
        "add_url": add_url,
        "delete_url": delete_url,
    }
    return render(request, 'edit.html', context)


# delete
def delete_sample(request, id):
    sample = Sample.objects.get(pk=id)
    title = 'Are you sure you want sample deleted?'
    pks = ('0', id)
    table = Sample.objects.filter(pk__in=pks)

    if request.method == "POST":
        if "del_btn" in request.POST:
            sample.delete()
            return HttpResponseRedirect('/archive/add_sample_single')
        if "no_btn" in request.POST:
            return HttpResponseRedirect('/archive/add_sample_single')
    else:
        context = {
            "title": title,
            "table": table,
        }
    return render(request, 'delete.html', context)


def delete_individual_multiple(request):
    add_url = 'add_individual_multiple'
    select_url = 'validate_individual_multiple'
    title = 'Individuals deleted.'
    table = 'n/a'
    pks = request.POST.getlist("selection")
    num_deleted = len(pks)
    individuals = Individual.objects.filter(pk__in=pks)
    individuals.delete()

    context = {
        "add_url": add_url,
        "select_url": select_url,
        "title": title,
        "table": table,
        "num_deleted": num_deleted,
    }
    return render(request, 'delete_multiple.html', context)


def delete_sample_multiple(request):
    add_url = 'add_sample_multiple'
    select_url = 'validate_sample_multiple'
    title = 'Samples deleted.'
    table = 'n/a'
    pks = request.POST.getlist("selection")
    num_deleted = len(pks)
    samples = Sample.objects.filter(pk__in=pks)
    samples.delete()

    context = {
        "add_url": add_url,
        "select_url": select_url,
        "title": title,
        "table": table,
        "num_deleted": num_deleted,
    }
    return render(request, 'delete_multiple.html', context)

# validate
def validate_individual_multiple(request):
    edit_url = 'edit_individual_multiple'
    delete_url = 'delete_individual_multiple'
    add_url = "add_individual_multiple"

    title = 'Select Individual'
    instructions = 'Please select Individual by clicking check box'
    qset = Individual.objects.all()
    filter = IndividualFilter(request.GET, queryset=qset)
    table = IndividualTable(filter.qs)

    context = {
        "button_type": 'buttons_2.html',
        "edit_url": edit_url,
        "delete_url": delete_url,
        "add_url": add_url,
        "title": title,
        "filter": filter,
        "instructions": instructions,
        "table": table,
    }
    context.update(validate_samples_button_context)
    return render(request, 'select.html', context)


def validate_sample_multiple(request):
    edit_url = 'edit_sample_multiple'
    delete_url = 'delete_sample_multiple'
    add_url = "add_sample_multiple"

    title = 'Select Samples'
    instructions = 'Please select samples by clicking check box'
    qset = Sample.objects.all()
    filter = SampleFilter(request.GET, queryset=qset)
    table = SampleTable(filter.qs)

    context = {
        "button_type": 'buttons_2.html',
        "edit_url": edit_url,
        "delete_url": delete_url,
        "add_url": add_url,
        "title": title,
        "filter": filter,
        "instructions": instructions,
        "table": table,
    }
    context.update(validate_samples_button_context)
    return render(request, 'select.html', context)


# edit
def edit_individual_multiple(request):
    title = 'Edit Individual'
    extras = 0
    IndividualFormSet = modelformset_factory(Individual, form=IndividualForm, extra=extras)
    pks = request.POST.getlist("selection")
    qset = Individual.objects.filter(pk__in=pks)
    formset = IndividualFormSet(queryset=qset)
    table = 'n/a'

    if "confirm_btn" in request.POST:
        formset = IndividualFormSet(request.POST)
        for form in formset:
            if form.is_valid():
                form.save()
        return HttpResponseRedirect('/archive/validate_individual_multiple')

    context = {
        "title": title,
        "formset": formset,
        "table": table,
    }
    return render(request, 'edit_multiple.html', context)


def edit_sample_multiple(request):
    title = 'Edit Samples'
    extras = 0
    SampleFormSet = modelformset_factory(Sample, form=SampleForm, extra=extras)
    pks = request.POST.getlist("selection")
    qset = Sample.objects.filter(pk__in=pks)
    formset = SampleFormSet(queryset=qset)
    table = 'n/a'

    if "confirm_btn" in request.POST:
        formset = SampleFormSet(request.POST)
        for form in formset:
            if form.is_valid():
                form.save()
        return HttpResponseRedirect('/archive/validate_sample_multiple')

    context = {
        "title": title,
        "formset": formset,
        "table": table,
    }
    return render(request, 'edit_multiple.html', context)


def model_form_upload(request):
    form2 = SampleForm()
    if request.method == 'POST':
        print('------------------form1--------------------')
        form1 = UploadFileForm(request.POST, request.FILES)
        if form1.is_valid():
            if "upload_btn" in request.POST:
                uploaded_file = request.FILES['myfile'].read()
                data = [row for row in csv.reader(uploaded_file.read().splitlines())]
                print(data)

        elif "submit_btn" in request.POST:
            print('-------------form1-------------------')
            print(form2)
        else:
            print('-----------no button---------------')

    else:
        print('----------failed----------')
        print(request.method)
        form1 = UploadFileForm()

    context = {
        'form1':form1,
        'form2':form2,
    }
    return render(request, 'model_form_upload.html', context)


# freezers and boxes
def add_freezer(request):
    title = 'Enter Freezer'
    form = FreezerForm(request.POST or None)
    form2 = 'n/a'
    table = FreezerTable(Freezer.objects.all())

    if form.is_valid():
        print('-----------------adding Freezer---------------')
        entered_freezer = form.save()
        sh = entered_freezer.shelves
        ra = entered_freezer.racks
        ro = entered_freezer.columns
        co = entered_freezer.rows
        print(sh)
        for i in range(sh):
            for j in range(ra):
                for k in range(ro):
                    for l in range(co):
                        freezerpos = FreezerPos(freezer=entered_freezer, shelf=i+1, rack=j+1, row=k+1, column=l+1)
                        print(freezerpos)
                        freezerpos.save()
        return HttpResponseRedirect('/archive/validate_freezer')

    context = {
        "button_type": 'buttons_2.html',
        "title": title,
        "form1": form,
        "form2": form2,
        "table": table,
    }
    context.update(add_freezer_button_context)
    return render(request, 'add_single.html', context)


def add_box(request):
    title = 'Enter Box'
    form = BoxForm(request.POST or None)
    form2 = 'n/a'
    table = BoxTable(Box.objects.all())

    if form.is_valid():
        print('-----------------adding Box---------------')
        entered_box = form.save()
        rows = entered_box.rows
        cols = entered_box.columns
        print(rows)
        for i in range(rows):
            row = ascii_uppercase[i]
            print(row)
            for j in range(cols):
                col = '%02d' % (j+1)
                print(col)
                boxpos = BoxPos(box=entered_box, row=row, column=col)
                print(boxpos)
                boxpos.save()
        #     for j in range(ra):
        #         for k in range(ro):
        #             for l in range(co):
        #                 freezerpos = FreezerPos(freezer=entered_sample, shelf=i+1, rack=j+1, row=k+1, column=l+1)
        #                 print(freezerpos)
        #                 freezerpos.save()
        return HttpResponseRedirect('/archive/validate_freezer')

    context = {
        "button_type": 'buttons_2.html',
        "title": title,
        "form1": form,
        "form2": form2,
        "table": table,
    }
    context.update(add_freezer_button_context)
    return render(request, 'add_single.html', context)

def validate_freezer(request):
    parent_url = 'validate_freezer'
    child_url = 'validate_sample_multiple'
    edit_url = 'edit_individual_multiple'
    delete_url = 'delete_individual_multiple'
    add_url = "add_freezer"

    title = 'Select Freezer'
    instructions = 'Please select Freezer by clicking check box'
    qset = Freezer.objects.all()
    filter = FreezerFilter(request.GET, queryset=qset)
    table = FreezerTable(filter.qs)

    context = {
        "parent_url": parent_url,
        "child_url": child_url,
        "edit_url": edit_url,
        "delete_url": delete_url,
        "add_url": add_url,
        "title": title,
        "filter": filter,
        "instructions": instructions,
        "table": table,
    }
    return render(request, 'select.html', context)

def validate_box(request):
    parent_url = 'validate_box'
    child_url = 'validate_sample_multiple'
    edit_url = 'edit_individual_multiple'
    delete_url = 'delete_individual_multiple'
    add_url = "add_freezer"

    title = 'Select Freezer'
    instructions = 'Please select Freezer by clicking check box'
    qset = Freezer.objects.all()
    filter = FreezerFilter(request.GET, queryset=qset)
    table = FreezerTable(filter.qs)

    context = {
        "parent_url": parent_url,
        "child_url": child_url,
        "edit_url": edit_url,
        "delete_url": delete_url,
        "add_url": add_url,
        "title": title,
        "filter": filter,
        "instructions": instructions,
        "table": table,
    }
    return render(request, 'select.html', context)

# testing

def export(request):
    sample_resource = SampleResource()
    dataset = sample_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="samples.xls"'
    return response


def import_data(request):
    form2 = SampleForm()
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['myfile']
            data = filehandle.get_records()
            print(data)
            form2 = SampleForm(initial=data[0])
            # return excel.make_response(filehandle.get_sheet(), "csv", file_name="download")
    else:
        form = UploadFileForm()

    context = {
        'form': form,
        'form2': form2,
    }

    return render(request, 'upload_form.html', context)


def simple_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save('temp.xlsx', myfile)
        uploaded_file_url = fs.url(filename)
        return render(request, 'simple_upload.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'simple_upload.html')


def upload(request):
    form2 = SampleForm()
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['myfile']
            data = filehandle.get_records()
            print(data)
            print(SampleForm(initial=data[0]))
            form2 = SampleForm(initial=data[0])
            # return excel.make_response(filehandle.get_sheet(), "csv", file_name="download")
    else:
        form = UploadFileForm()

    context = {
        'form': form,
        'form2': form2,
    }

    return render(request, 'upload_form.html', context)


